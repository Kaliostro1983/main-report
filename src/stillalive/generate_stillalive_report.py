from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from src.armorkit.data_loader import load_config, load_inputs
from src.armorkit.normalize_freq import normalize_frequency_column
from src.armorkit.dates import parse_period_from_filename, format_for_filename


def _get_observed_frequencies(reference_df: pd.DataFrame) -> pd.DataFrame:
    """
    Вибирає частоти, які перебувають на спостереженні (Статус == 'Спостерігається').

    Повертає датафрейм з колонками:
    - 'Частота'
    - 'Маска_3'
    - 'Підрозділ' (якщо є у reference_df)
    - 'Дешифрування' (якщо є у reference_df)
    """
    observed = reference_df[reference_df["Статус"] == "Спостерігається"].copy()

    # Додаємо нові поля, якщо їх немає в довіднику — створюємо порожні
    if "Підрозділ" not in observed.columns:
        observed["Підрозділ"] = ""
    if "Дешифрування" not in observed.columns:
        observed["Дешифрування"] = ""

    observed = observed[["Частота", "Маска_3", "Підрозділ", "Дешифрування"]].drop_duplicates(
        subset=["Частота"], keep="first"
    )
    observed["Частота"] = observed["Частота"].astype(str)
    observed["Підрозділ"] = observed["Підрозділ"].astype(str).fillna("")
    observed["Дешифрування"] = observed["Дешифрування"].astype(str).fillna("")
    return observed


def _prepare_daily_counts(intercepts_df: pd.DataFrame,
                          observed_freqs_df: pd.DataFrame) -> pd.DataFrame:
    """
    Формує підсумковий датафрейм у wide-форматі:
    - "Частота"
    - "Маска"
    - "Підрозділ"
    - "Дешифрування"
    - далі по одній колонці на кожен день періоду (формат "dd.mm")
      Значення = кількість перехоплень цієї частоти у цю дату.

    ВАЖЛИВО: включаємо всі частоти зі статусом "Спостерігається",
    навіть якщо перехоплень по них 0.
    """

    work = intercepts_df.copy()
    work["Частота"] = work["Частота"].astype(str)

    if "Дата" not in work.columns:
        raise ValueError("У перехопленнях немає колонки 'Дата', не можу побудувати добову активність.")

    # Нормалізація дати
    work["_dt"] = pd.to_datetime(work["Дата"], errors="coerce")
    work = work.dropna(subset=["_dt"])
    work["_day_label"] = work["_dt"].dt.strftime("%d.%m")

    # Лічильник перехоплень по (Частота, день)
    counts = (
        work.groupby(["Частота", "_day_label"])
        .size()
        .reset_index(name="count")
    )

    # Всі унікальні дні періоду
    all_days = (
        counts["_day_label"]
        .drop_duplicates()
        .sort_values(key=lambda s: s.apply(lambda x: (int(x.split(".")[1]), int(x.split(".")[0]))))
        .tolist()
    )

    # Повний список частот зі статусом "Спостерігається"
    freqs_df = observed_freqs_df.copy()
    freqs_df["Частота"] = freqs_df["Частота"].astype(str)
    freqs_df = freqs_df.drop_duplicates(subset=["Частота"], keep="first")

    # Повна комбінація (частота × день)
    freq_list = freqs_df["Частота"].tolist()
    cartesian = (
        pd.MultiIndex.from_product(
            [freq_list, all_days],
            names=["Частота", "_day_label"]
        ).to_frame(index=False)
    )

    cartesian = cartesian.merge(counts, on=["Частота", "_day_label"], how="left")
    cartesian["count"] = cartesian["count"].fillna(0).astype(int)

    # Wide-формат
    pivot = cartesian.pivot_table(
        index="Частота",
        columns="_day_label",
        values="count",
        fill_value=0,
        aggfunc="sum"
    ).reset_index()

    # Мапи з довідника частот
    mask_map = freqs_df.set_index("Частота")["Маска_3"].to_dict()
    unit_map = freqs_df.set_index("Частота")["Підрозділ"].to_dict() if "Підрозділ" in freqs_df.columns else {}
    dec_map = freqs_df.set_index("Частота")["Дешифрування"].to_dict() if "Дешифрування" in freqs_df.columns else {}

    # Додаємо службові колонки зліва
    pivot.insert(1, "Маска", pivot["Частота"].map(mask_map).fillna(""))
    pivot.insert(2, "Підрозділ", pivot["Частота"].map(unit_map).fillna(""))
    pivot.insert(3, "Дешифрування", pivot["Частота"].map(dec_map).fillna(""))

    final_cols = ["Частота", "Маска", "Підрозділ", "Дешифрування"] + all_days
    pivot = pivot[final_cols]

    return pivot


def _export_to_xlsx(df: pd.DataFrame, period_start: str, period_end: str, output_dir: Path) -> Path:
    """
    Записує df у Excel з форматуванням:
    - Заголовки колонок жирним.
    - Колонки: "Частота", "Маска", "Підрозділ", "Дешифрування" — жирним у всіх рядках.
    - Комірки дат: 0 -> червоний фон, >0 -> зелений фон.
    - Якщо "Дешифрування" != 'так' (регістр/пробіли ігноруємо), то
      комірки "Частота", "Маска", "Підрозділ", "Дешифрування" — червоний фон.
    """

    start_s = format_for_filename(period_start)
    end_s = format_for_filename(period_end)
    filename = f"Активність радіомереж ({start_s} - {end_s}).xlsx"
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / filename

    wb = Workbook()
    ws = wb.active
    ws.title = "still_alive"

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    n_rows = ws.max_row
    n_cols = ws.max_column

    bold_font = Font(bold=True)
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Заголовки
    for c in range(1, n_cols + 1):
        ws.cell(row=1, column=c).font = bold_font

    # Службові колонки (1..4) — жирним
    meta_cols = 4 if n_cols >= 4 else n_cols
    for r in range(1, n_rows + 1):
        for c in range(1, meta_cols + 1):
            ws.cell(row=r, column=c).font = bold_font

    # Якщо дешифрування НЕ "так" — підсвічуємо 1..4 червоним
    # (значення беремо з колонки 4)
    if n_cols >= 4:
        for r in range(2, n_rows + 1):
            dec_val = ws.cell(row=r, column=4).value
            dec_norm = str(dec_val).strip().lower() if dec_val is not None else ""
            if dec_norm != "так":
                for c in range(1, 5):
                    ws.cell(row=r, column=c).fill = red_fill

    # Підсвітка значень по днях (починаючи з 5-ї колонки)
    day_start_col = 5 if n_cols >= 5 else (meta_cols + 1)
    for r in range(2, n_rows + 1):
        for c in range(day_start_col, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            try:
                num = int(cell.value)
                if num == 0:
                    cell.fill = red_fill
                else:
                    cell.fill = green_fill
            except Exception:
                continue

    wb.save(out_path)
    return out_path


def build_stillalive_report(config_path: str) -> Path:
    """
    Основний пайплайн:
    1. Завантажує конфіг і дані.
    2. Визначає період звіту з назви файлу.
    3. Нормалізує частоти у перехопленнях.
    4. Вибирає частоти зі статусом "Спостерігається".
    5. Формує підсумковий датафрейм по днях.
    6. Зберігає Excel у build/.
    """

    cfg = load_config(config_path)
    li = load_inputs(config_path)

    period_start, period_end = parse_period_from_filename(li.report_path)
    normalize_frequency_column(li.intercepts_df, li.reference_df, li.masks_df)

    observed_freqs_df = _get_observed_frequencies(li.reference_df)
    summary_df = _prepare_daily_counts(li.intercepts_df, observed_freqs_df)

    out_dir = Path(getattr(cfg.paths, "output_dir", "build"))
    out_path = _export_to_xlsx(summary_df, period_start, period_end, out_dir)

    return out_path
